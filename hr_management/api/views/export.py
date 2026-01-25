"""数据导出 API 视图"""
from io import BytesIO
from datetime import datetime

from django.http import HttpResponse
from rest_framework import permissions, views
from rest_framework.decorators import api_view, permission_classes
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ...models import Employee, Attendance, LeaveRequest, SalaryRecord


def _previous_year_month(dt):
    if dt.month == 1:
        return dt.year - 1, 12
    return dt.year, dt.month - 1


def style_header(ws, headers):
    """设置表头样式"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def create_excel_response(wb, filename):
    """创建 Excel 下载响应"""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAdminUser])
def export_employees(request):
    """导出员工列表 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "员工列表"
    
    headers = ['员工编号', '姓名', '性别', '部门', '职位', '手机号', '邮箱', '入职日期', '状态']
    style_header(ws, headers)
    
    employees = Employee.objects.select_related('department', 'position').all()
    
    for row, emp in enumerate(employees, 2):
        ws.cell(row=row, column=1, value=emp.employee_id)
        ws.cell(row=row, column=2, value=emp.name)
        ws.cell(row=row, column=3, value=dict(Employee.GENDER_CHOICES).get(emp.gender, ''))
        ws.cell(row=row, column=4, value=emp.department.name if emp.department else '')
        ws.cell(row=row, column=5, value=emp.position.name if emp.position else '')
        ws.cell(row=row, column=6, value=emp.phone)
        ws.cell(row=row, column=7, value=emp.email)
        ws.cell(row=row, column=8, value=emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '')
        ws.cell(row=row, column=9, value='在职' if emp.is_active else '离职')
    
    # 调整列宽
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    filename = f"员工列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@api_view(['GET'])
@permission_classes([permissions.IsAdminUser])
def export_salaries(request):
    """导出薪资记录 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "薪资记录"
    
    headers = ['员工编号', '姓名', '部门', '年份', '月份', '基本工资', '奖金', '津贴', '实发工资', '发放状态']
    style_header(ws, headers)
    
    # 支持筛选参数
    ids = request.query_params.get('ids')
    year = request.query_params.get('year')
    month = request.query_params.get('month')
    
    qs = SalaryRecord.objects.select_related('employee', 'employee__department').all()
    
    # 优先按 ids 筛选
    if ids:
        id_list = [int(i) for i in ids.split(',') if i.strip().isdigit()]
        if id_list:
            qs = qs.filter(id__in=id_list)
    else:
        if year:
            qs = qs.filter(year=year)
        if month:
            qs = qs.filter(month=month)
    
    for row, sal in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=sal.employee.employee_id)
        ws.cell(row=row, column=2, value=sal.employee.name)
        ws.cell(row=row, column=3, value=sal.employee.department.name if sal.employee.department else '')
        ws.cell(row=row, column=4, value=sal.year)
        ws.cell(row=row, column=5, value=sal.month)
        ws.cell(row=row, column=6, value=float(sal.basic_salary))
        ws.cell(row=row, column=7, value=float(sal.bonus))
        ws.cell(row=row, column=8, value=float(sal.allowance))
        ws.cell(row=row, column=9, value=float(sal.net_salary))
        ws.cell(row=row, column=10, value='已发放' if sal.paid else '未发放')
    
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 12
    
    filename = f"薪资记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_my_salary_slip(request):
    """普通用户导出本人某月工资条（Excel）

    - 仅允许导出当前登录用户关联员工的工资条
    - 支持 year/month 参数；缺省则默认上个月
    """
    try:
        emp = Employee.objects.select_related('department').get(user=request.user)
    except Employee.DoesNotExist:
        return HttpResponse('当前账户未关联员工，无法导出工资条', status=400)

    year = request.query_params.get('year')
    month = request.query_params.get('month')
    if not year or not month:
        now = timezone.localtime()
        py, pm = _previous_year_month(now)
        year = year or str(py)
        month = month or str(pm)

    try:
        year_i = int(year)
        month_i = int(month)
        if month_i < 1 or month_i > 12:
            raise ValueError('month out of range')
    except Exception:
        return HttpResponse('year/month 参数不合法', status=400)

    rec = SalaryRecord.objects.filter(employee=emp, year=year_i, month=month_i).first()
    if not rec:
        return HttpResponse('未找到该月工资记录', status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "工资条"

    headers = ['员工编号', '姓名', '部门', '年份', '月份', '基本工资', '奖金', '津贴', '实发工资', '发放状态', '发放时间']
    style_header(ws, headers)

    ws.cell(row=2, column=1, value=emp.employee_id)
    ws.cell(row=2, column=2, value=emp.name)
    ws.cell(row=2, column=3, value=emp.department.name if emp.department else '')
    ws.cell(row=2, column=4, value=rec.year)
    ws.cell(row=2, column=5, value=rec.month)
    ws.cell(row=2, column=6, value=float(rec.basic_salary))
    ws.cell(row=2, column=7, value=float(rec.bonus))
    ws.cell(row=2, column=8, value=float(rec.allowance))
    ws.cell(row=2, column=9, value=float(rec.net_salary))
    ws.cell(row=2, column=10, value='已发放' if rec.paid else '未发放')
    ws.cell(row=2, column=11, value=rec.paid_at.strftime('%Y-%m-%d %H:%M:%S') if rec.paid_at else '')

    for col in range(1, 12):
        ws.column_dimensions[chr(64 + col)].width = 14

    filename = f"工资条_{emp.employee_id}_{year_i}{str(month_i).zfill(2)}.xlsx"
    return create_excel_response(wb, filename)


@api_view(['GET'])
@permission_classes([permissions.IsAdminUser])
def export_attendance(request):
    """导出考勤记录 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤记录"
    
    headers = ['员工编号', '姓名', '部门', '日期', '上班时间', '下班时间', '考勤类型', '备注']
    style_header(ws, headers)
    
    # 支持日期筛选
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')
    
    qs = Attendance.objects.select_related('employee', 'employee__department').all()
    if start_date:
        qs = qs.filter(date__gte=start_date)
    if end_date:
        qs = qs.filter(date__lte=end_date)
    
    type_map = dict(Attendance.ATTENDANCE_TYPE_CHOICES)
    
    for row, att in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=att.employee.employee_id)
        ws.cell(row=row, column=2, value=att.employee.name)
        ws.cell(row=row, column=3, value=att.employee.department.name if att.employee.department else '')
        ws.cell(row=row, column=4, value=att.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=att.check_in_time.strftime('%H:%M') if att.check_in_time else '')
        ws.cell(row=row, column=6, value=att.check_out_time.strftime('%H:%M') if att.check_out_time else '')
        ws.cell(row=row, column=7, value=type_map.get(att.attendance_type, att.attendance_type))
        ws.cell(row=row, column=8, value=att.notes)
    
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 14
    
    filename = f"考勤记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@api_view(['GET'])
@permission_classes([permissions.IsAdminUser])
def export_leaves(request):
    """导出请假记录 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "请假记录"
    
    headers = ['员工编号', '姓名', '部门', '请假类型', '开始日期', '结束日期', '天数', '状态', '申请时间']
    style_header(ws, headers)
    
    qs = LeaveRequest.objects.select_related('employee', 'employee__department').all()
    
    type_map = dict(LeaveRequest.LEAVE_TYPE_CHOICES)
    status_map = dict(LeaveRequest.STATUS_CHOICES)
    
    for row, lv in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=lv.employee.employee_id)
        ws.cell(row=row, column=2, value=lv.employee.name)
        ws.cell(row=row, column=3, value=lv.employee.department.name if lv.employee.department else '')
        ws.cell(row=row, column=4, value=type_map.get(lv.leave_type, lv.leave_type))
        ws.cell(row=row, column=5, value=lv.start_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=6, value=lv.end_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=7, value=lv.days)
        ws.cell(row=row, column=8, value=status_map.get(lv.status, lv.status))
        ws.cell(row=row, column=9, value=lv.created_at.strftime('%Y-%m-%d %H:%M'))
    
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 14
    
    filename = f"请假记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return create_excel_response(wb, filename)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_salary_template(request):
    """导出薪资导入模板 Excel"""
    from openpyxl.styles import NamedStyle, Protection
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "薪资数据"
    
    headers = ['员工姓名', '年份', '月份', '基本工资', '奖金', '津贴']
    style_header(ws, headers)
    
    # 示例数据行样式 - 浅灰色背景，斜体
    example_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    example_font = Font(color="999999", italic=True)
    example_align = Alignment(horizontal="center", vertical="center")
    
    example_data = ['张三', datetime.now().year, datetime.now().month, 8000, 500, 300]
    for col, val in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = example_font
        cell.fill = example_fill
        cell.alignment = example_align
    
    # 设置列宽
    column_widths = [15, 10, 8, 14, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # ========== 创建说明 Sheet ==========
    ws_help = wb.create_sheet(title="填写说明")
    
    # 说明标题
    title_font = Font(bold=True, size=14, color="4472C4")
    ws_help.cell(row=1, column=1, value="薪资导入模板 - 填写说明").font = title_font
    ws_help.merge_cells('A1:D1')
    
    # 说明内容
    instructions = [
        "",
        "【使用步骤】",
        "1. 在「薪资数据」工作表中填写数据",
        "2. 第2行为示例数据（灰色斜体），导入前请删除",
        "3. 从第2行开始填写实际薪资数据",
        "",
        "【字段说明】",
        "• 员工姓名：必填，必须与系统中的员工姓名完全一致",
        "• 年份：必填，四位数字，如 2026",
        "• 月份：必填，1-12 的数字",
        "• 基本工资：必填，数字格式，不含货币符号",
        "• 奖金：选填，数字格式，默认为 0",
        "• 津贴：选填，数字格式，默认为 0",
        "",
        "【注意事项】",
        "• 同一员工同一月份只能有一条薪资记录",
        "• 金额支持小数，如 8500.50",
        "• 请勿修改表头行",
    ]
    
    help_font = Font(size=11)
    section_font = Font(bold=True, size=11, color="333333")
    
    for i, text in enumerate(instructions, 2):
        cell = ws_help.cell(row=i, column=1, value=text)
        if text.startswith("【") and text.endswith("】"):
            cell.font = section_font
        else:
            cell.font = help_font
    
    ws_help.column_dimensions['A'].width = 50
    
    # 激活第一个工作表
    wb.active = ws
    
    filename = "薪资导入模板.xlsx"
    return create_excel_response(wb, filename)
