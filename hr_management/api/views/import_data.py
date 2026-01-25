"""批量导入 API 视图"""
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.db import transaction
from rest_framework import permissions, views
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
import openpyxl

from ...models import Employee, Department, Position, Attendance, SalaryRecord
from ...utils import log_event, api_success, api_error, get_client_ip


User = get_user_model()


class BaseImportAPIView(views.APIView):
    """批量导入基类"""
    permission_classes = [permissions.IsAdminUser]
    parser_classes = [MultiPartParser]
    model_name = ''  # 子类设置
    
    def parse_excel(self, file):
        """解析 Excel 文件，返回行列表"""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return None, '文件为空或只有表头'
            headers = [str(h).strip() if h else '' for h in rows[0]]
            data_rows = rows[1:]
            return {'headers': headers, 'data': data_rows}, None
        except Exception as e:
            return None, f'解析 Excel 失败: {str(e)}'
    
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response(api_error('请上传文件'), status=400)
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(api_error('仅支持 Excel 文件(.xlsx, .xls)'), status=400)
        
        parsed, err = self.parse_excel(file)
        if err:
            return Response(api_error(err), status=400)
        
        result, errors = self.import_data(parsed, request)
        
        log_event(
            user=request.user,
            action=f'批量导入{self.model_name}',
            detail=f'成功 {result["success"]} 条，失败 {result["failed"]} 条',
            ip=get_client_ip(request)
        )
        
        return Response(api_success(result))
    
    def import_data(self, parsed, request):
        """子类实现具体导入逻辑"""
        raise NotImplementedError


class EmployeeImportAPIView(BaseImportAPIView):
    """员工批量导入"""
    model_name = '员工'
    
    # 表头映射
    FIELD_MAP = {
        '工号': 'employee_id',
        '员工工号': 'employee_id',
        '姓名': 'name',
        '员工姓名': 'name',
        '性别': 'gender',
        '手机号': 'phone',
        '联系电话': 'phone',
        '邮箱': 'email',
        '电子邮箱': 'email',
        '部门': 'department',
        '所属部门': 'department',
        '职位': 'position',
        '岗位': 'position',
        '入职日期': 'hire_date',
        '基本工资': 'salary',
        '薪资': 'salary',
        '身份证号': 'id_card',
        '地址': 'address',
        '家庭住址': 'address',
        '出生日期': 'birth_date',
        '紧急联系人': 'emergency_contact',
        '紧急联系电话': 'emergency_phone',
        '备注': 'notes',
    }
    
    def import_data(self, parsed, request):
        headers = parsed['headers']
        data_rows = parsed['data']
        
        # 构建列索引映射
        col_map = {}
        for i, h in enumerate(headers):
            field = self.FIELD_MAP.get(h)
            if field:
                col_map[field] = i
        
        # 检查必填列
        required = ['employee_id', 'name']
        missing = [f for f in required if f not in col_map]
        if missing:
            return {'success': 0, 'failed': len(data_rows), 'errors': [f'缺少必填列: {missing}']}
        
        # 预加载部门和职位
        depts = {d.name: d for d in Department.objects.all()}
        positions = {p.name: p for p in Position.objects.all()}
        
        success = 0
        failed = 0
        errors = []
        
        for row_idx, row in enumerate(data_rows, start=2):
            try:
                with transaction.atomic():
                    data = {field: row[idx] if idx < len(row) else None for field, idx in col_map.items()}
                    
                    emp_id = str(data.get('employee_id', '')).strip()
                    name = str(data.get('name', '')).strip()
                    
                    if not emp_id or not name:
                        errors.append(f'第 {row_idx} 行: 工号或姓名为空')
                        failed += 1
                        continue
                    
                    # 查找或创建员工
                    emp, created = Employee.objects.get_or_create(
                        employee_id=emp_id,
                        defaults={'name': name}
                    )
                    
                    # 更新字段
                    emp.name = name
                    
                    if data.get('gender'):
                        gender_map = {'男': 'male', '女': 'female', 'male': 'male', 'female': 'female'}
                        emp.gender = gender_map.get(str(data['gender']).strip().lower(), emp.gender)
                    
                    if data.get('phone'):
                        emp.phone = str(data['phone']).strip()
                    
                    if data.get('email'):
                        emp.email = str(data['email']).strip()
                    
                    if data.get('department'):
                        dept_name = str(data['department']).strip()
                        if dept_name in depts:
                            emp.department = depts[dept_name]
                        else:
                            errors.append(f'第 {row_idx} 行: 部门 "{dept_name}" 不存在')
                    
                    if data.get('position'):
                        pos_name = str(data['position']).strip()
                        if pos_name in positions:
                            emp.position = positions[pos_name]
                        else:
                            errors.append(f'第 {row_idx} 行: 职位 "{pos_name}" 不存在')
                    
                    if data.get('hire_date'):
                        try:
                            hire_date = data['hire_date']
                            if isinstance(hire_date, datetime):
                                emp.hire_date = hire_date.date()
                            else:
                                emp.hire_date = datetime.strptime(str(hire_date).strip(), '%Y-%m-%d').date()
                        except:
                            pass
                    
                    if data.get('salary'):
                        try:
                            emp.salary = Decimal(str(data['salary']))
                        except:
                            pass
                    
                    if data.get('id_card'):
                        emp.id_card = str(data['id_card']).strip()
                    
                    if data.get('address'):
                        emp.address = str(data['address']).strip()
                    
                    if data.get('birth_date'):
                        try:
                            bd = data['birth_date']
                            if isinstance(bd, datetime):
                                emp.birth_date = bd.date()
                            else:
                                emp.birth_date = datetime.strptime(str(bd).strip(), '%Y-%m-%d').date()
                        except:
                            pass
                    
                    if data.get('emergency_contact'):
                        emp.emergency_contact = str(data['emergency_contact']).strip()
                    
                    if data.get('emergency_phone'):
                        emp.emergency_phone = str(data['emergency_phone']).strip()
                    
                    if data.get('notes'):
                        emp.notes = str(data['notes']).strip()
                    
                    emp.save()
                    success += 1
                    
            except Exception as e:
                errors.append(f'第 {row_idx} 行: {str(e)}')
                failed += 1
        
        return {'success': success, 'failed': failed, 'errors': errors[:50]}  # 最多返回50条错误


class AttendanceImportAPIView(BaseImportAPIView):
    """考勤批量导入"""
    model_name = '考勤'
    
    FIELD_MAP = {
        '工号': 'employee_id',
        '员工工号': 'employee_id',
        '日期': 'date',
        '考勤日期': 'date',
        '类型': 'attendance_type',
        '考勤类型': 'attendance_type',
        '上班时间': 'check_in',
        '签到时间': 'check_in',
        '下班时间': 'check_out',
        '签退时间': 'check_out',
        '备注': 'notes',
    }
    
    TYPE_MAP = {
        '正常': 'normal',
        '迟到': 'late',
        '早退': 'early_leave',
        '缺勤': 'absent',
        '请假': 'leave',
        '加班': 'overtime',
        '出差': 'business_trip',
        'normal': 'normal',
        'late': 'late',
        'early_leave': 'early_leave',
        'absent': 'absent',
        'leave': 'leave',
        'overtime': 'overtime',
        'business_trip': 'business_trip',
    }
    
    def import_data(self, parsed, request):
        headers = parsed['headers']
        data_rows = parsed['data']
        
        col_map = {}
        for i, h in enumerate(headers):
            field = self.FIELD_MAP.get(h)
            if field:
                col_map[field] = i
        
        required = ['employee_id', 'date']
        missing = [f for f in required if f not in col_map]
        if missing:
            return {'success': 0, 'failed': len(data_rows), 'errors': [f'缺少必填列: {missing}']}
        
        # 预加载员工
        employees = {e.employee_id: e for e in Employee.objects.all()}
        
        success = 0
        failed = 0
        errors = []
        
        for row_idx, row in enumerate(data_rows, start=2):
            try:
                with transaction.atomic():
                    data = {field: row[idx] if idx < len(row) else None for field, idx in col_map.items()}
                    
                    emp_id = str(data.get('employee_id', '')).strip()
                    if emp_id not in employees:
                        errors.append(f'第 {row_idx} 行: 工号 "{emp_id}" 不存在')
                        failed += 1
                        continue
                    
                    employee = employees[emp_id]
                    
                    # 解析日期
                    date_val = data.get('date')
                    if isinstance(date_val, datetime):
                        date_obj = date_val.date()
                    else:
                        try:
                            date_obj = datetime.strptime(str(date_val).strip(), '%Y-%m-%d').date()
                        except:
                            errors.append(f'第 {row_idx} 行: 日期格式错误')
                            failed += 1
                            continue
                    
                    # 考勤类型
                    att_type = 'normal'
                    if data.get('attendance_type'):
                        att_type = self.TYPE_MAP.get(str(data['attendance_type']).strip(), 'normal')
                    
                    # 解析时间
                    check_in = None
                    check_out = None
                    if data.get('check_in'):
                        ci = data['check_in']
                        if isinstance(ci, datetime):
                            check_in = ci.time()
                        else:
                            try:
                                check_in = datetime.strptime(str(ci).strip(), '%H:%M:%S').time()
                            except:
                                try:
                                    check_in = datetime.strptime(str(ci).strip(), '%H:%M').time()
                                except:
                                    pass
                    
                    if data.get('check_out'):
                        co = data['check_out']
                        if isinstance(co, datetime):
                            check_out = co.time()
                        else:
                            try:
                                check_out = datetime.strptime(str(co).strip(), '%H:%M:%S').time()
                            except:
                                try:
                                    check_out = datetime.strptime(str(co).strip(), '%H:%M').time()
                                except:
                                    pass
                    
                    # 更新或创建
                    att, created = Attendance.objects.update_or_create(
                        employee=employee,
                        date=date_obj,
                        defaults={
                            'attendance_type': att_type,
                            'check_in': check_in,
                            'check_out': check_out,
                            'notes': str(data.get('notes', '') or '').strip(),
                        }
                    )
                    success += 1
                    
            except Exception as e:
                errors.append(f'第 {row_idx} 行: {str(e)}')
                failed += 1
        
        return {'success': success, 'failed': failed, 'errors': errors[:50]}


class SalaryImportAPIView(BaseImportAPIView):
    """薪资批量导入"""
    model_name = '薪资'
    
    FIELD_MAP = {
        '工号': 'employee_id',
        '员工工号': 'employee_id',
        '年份': 'year',
        '月份': 'month',
        '基本工资': 'base_salary',
        '奖金': 'bonus',
        '津贴': 'allowance',
        '备注': 'notes',
    }
    
    def import_data(self, parsed, request):
        headers = parsed['headers']
        data_rows = parsed['data']
        
        col_map = {}
        for i, h in enumerate(headers):
            field = self.FIELD_MAP.get(h)
            if field:
                col_map[field] = i
        
        required = ['employee_id', 'year', 'month', 'base_salary']
        missing = [f for f in required if f not in col_map]
        if missing:
            return {'success': 0, 'failed': len(data_rows), 'errors': [f'缺少必填列: {missing}']}
        
        employees = {e.employee_id: e for e in Employee.objects.all()}
        
        success = 0
        failed = 0
        errors = []
        
        for row_idx, row in enumerate(data_rows, start=2):
            try:
                with transaction.atomic():
                    data = {field: row[idx] if idx < len(row) else None for field, idx in col_map.items()}
                    
                    emp_id = str(data.get('employee_id', '')).strip()
                    if emp_id not in employees:
                        errors.append(f'第 {row_idx} 行: 工号 "{emp_id}" 不存在')
                        failed += 1
                        continue
                    
                    employee = employees[emp_id]
                    
                    try:
                        year = int(data['year'])
                        month = int(data['month'])
                        if not (1 <= month <= 12):
                            raise ValueError('月份超出范围')
                    except:
                        errors.append(f'第 {row_idx} 行: 年份或月份格式错误')
                        failed += 1
                        continue
                    
                    try:
                        base_salary = Decimal(str(data['base_salary']))
                    except:
                        errors.append(f'第 {row_idx} 行: 基本工资格式错误')
                        failed += 1
                        continue
                    
                    bonus = Decimal('0')
                    allowance = Decimal('0')
                    
                    if data.get('bonus'):
                        try:
                            bonus = Decimal(str(data['bonus']))
                        except:
                            pass
                    
                    if data.get('allowance'):
                        try:
                            allowance = Decimal(str(data['allowance']))
                        except:
                            pass
                    
                    actual_salary = base_salary + bonus + allowance
                    
                    salary, created = SalaryRecord.objects.update_or_create(
                        employee=employee,
                        year=year,
                        month=month,
                        defaults={
                            'base_salary': base_salary,
                            'bonus': bonus,
                            'allowance': allowance,
                            'actual_salary': actual_salary,
                            'notes': str(data.get('notes', '') or '').strip(),
                        }
                    )
                    success += 1
                    
            except Exception as e:
                errors.append(f'第 {row_idx} 行: {str(e)}')
                failed += 1
        
        return {'success': success, 'failed': failed, 'errors': errors[:50]}


class ImportTemplateAPIView(views.APIView):
    """下载导入模板"""
    permission_classes = [permissions.IsAdminUser]
    
    TEMPLATES = {
        'employee': {
            'filename': '员工导入模板.xlsx',
            'headers': ['工号', '姓名', '性别', '手机号', '邮箱', '部门', '职位', '入职日期', '基本工资', '身份证号', '地址', '出生日期', '紧急联系人', '紧急联系电话', '备注'],
            'example': ['EMP001', '张三', '男', '13800138000', 'zhangsan@example.com', '技术部', '工程师', '2024-01-01', '10000', '110101199001011234', '北京市朝阳区', '1990-01-01', '李四', '13900139000', ''],
        },
        'attendance': {
            'filename': '考勤导入模板.xlsx',
            'headers': ['工号', '日期', '考勤类型', '上班时间', '下班时间', '备注'],
            'example': ['EMP001', '2024-01-01', '正常', '09:00', '18:00', ''],
        },
        'salary': {
            'filename': '薪资导入模板.xlsx',
            'headers': ['工号', '年份', '月份', '基本工资', '奖金', '津贴', '备注'],
            'example': ['EMP001', '2024', '1', '10000', '2000', '500', ''],
        },
    }
    
    def get(self, request, template_type):
        if template_type not in self.TEMPLATES:
            return Response(api_error('模板类型不存在'), status=404)
        
        template = self.TEMPLATES[template_type]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '导入数据'
        
        # 写入表头
        for col, header in enumerate(template['headers'], start=1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入示例数据
        for col, value in enumerate(template['example'], start=1):
            ws.cell(row=2, column=col, value=value)
        
        # 保存到内存
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{template["filename"]}"'
        wb.save(response)
        return response
